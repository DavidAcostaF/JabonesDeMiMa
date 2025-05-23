from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Font, PatternFill, Side, Alignment


class ReportGenerator:

    def init_excel(delete_default_sheet=True):
        """
        Initialize an Excel workbook and return it as a BytesIO object.

        Parameters:
        delete_default_sheet (bool): If True, the default sheet will be removed from the workbook.
        """
        # Initialize the Excel workbook
        output = BytesIO()
        workbook = Workbook()
        if delete_default_sheet:
            # Remove the default sheet
            workbook.remove(workbook.active)
        return output, workbook

    def save_excel(output, workbook):
        """
        Save the Excel workbook to a file.

        Parameters:
        output (BytesIO): The BytesIO object containing the workbook.
        workbook (openpyxl.Workbook): The workbook to save.

        return (int): The position of the BytesIO object.
        """
        # Save the workbook to the output BytesIO object
        workbook.save(output)
        # Seek to the beginning of the BytesIO object
        return output.seek(0)

    def fill_worksheet(
        worksheet,
        data,
        columns,
        center = False,
        fill_y=PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
        fill_z=PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        cell_font=Font(color="000000", name="Arial"),
        border=Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
    ):
        """
        Fill an Excel worksheet with data.

        Parameters:
        worksheet (openpyxl.Worksheet): The worksheet to fill.
        data (list): A list of dictionaries containing the data to fill the worksheet with.
        columns (list): A list of column titles to add.
        center (boolean): if true: center the columns 
        """
        row_start = 3
        if isinstance(columns, dict):
            row_start = 4
            columns = [title for titles in columns.values() for title in titles]
        # Add the data to the worksheet
        for row_idx, row_data in enumerate(data, start=row_start):
            row_fill = fill_y if row_idx % 2 == 0 else fill_z
            for col_idx in range(len(columns)):
                cell = worksheet.cell(row=row_idx, column=col_idx + 2, value=row_data[col_idx])
                cell.fill = row_fill
                cell.font = cell_font  
                cell.border = border
                if center:
                    cell.alignment = Alignment(horizontal="center")


    def fill_column_titles(
        worksheet,
        columns,
        column_width=20,
        center = False,
        header_fill=PatternFill(start_color="004987", end_color="004987", fill_type="solid"),
        header_font=Font(bold=True, color="FFFFFF", name="Arial Bold"),
        border=Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
    ):
        """
        Add column titles to an Excel worksheet.

        Parameters:
        worksheet (openpyxl.Worksheet): The worksheet to add the titles to.
        columns (list/dict): A list of column titles to add.
        center (boolean): if true: center the columns 
        """
        def apply_header_styles(cell,center):
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            if center:
                cell.alignment = Alignment(horizontal="center")

        if isinstance(columns, dict):
            col_idx = 2
            # Add headers and merge cells for each key in the dictionary
            for key, titles in columns.items():
                # Merge cells for the header
                start_col = get_column_letter(col_idx)
                end_col = get_column_letter(col_idx + len(titles) - 1)
                worksheet.merge_cells(f"{start_col}2:{end_col}2")
                worksheet.merge_cells(f"{start_col}2:{end_col}2")
                cell = worksheet.cell(row=2, column=col_idx, value=key)
                apply_header_styles(cell, center=True)

                # Add the titles in the next row
                for idx, title in enumerate(titles, start=col_idx):
                    cell = worksheet.cell(row=3, column=idx, value=title)
                    apply_header_styles(cell, center=center)
                    worksheet.column_dimensions[get_column_letter(idx)].width = column_width

                col_idx += len(titles)
        elif isinstance(columns, list):
            # Add the column titles to the worksheet (starting from the second row and second column)
            for idx, title in enumerate(columns, start=1):
                cell = worksheet.cell(row=2, column=idx + 1, value=title)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = column_width